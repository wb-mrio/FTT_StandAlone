# -*- coding: utf-8 -*-
"""
Created on Thu Mar  7 15:32:07 2024

@author: Femke Nijsse
"""

from openpyxl import load_workbook

# Load the workbook
workbook_path = 'FTT-P-24x71_2022_S0.xlsx'
workbook = load_workbook(workbook_path)

# Sheets to ignore
sheets_to_ignore = ["Info", 'RERY', 'TITLES', 'MPRT', "MWDD", "MWDT", "MEWB", "MEWW", "MJET", "MCCS"]


for sheet_name in workbook.sheetnames:
    if sheet_name not in sheets_to_ignore:
        sheet = workbook[sheet_name]
        max_col = sheet.max_column
        
        # Parameters
        max_row = sheet.max_row
        max_col = sheet.max_column

        # 1. Remove the content from column C and further for specific rows
        for i in range(80):  # Assuming "and so forth" to cover about 80 iterations
            rows_to_clear = [16 + i*36, 17 + i*36]
            rows_to_clear = [row for row in rows_to_clear if row <= max_row]
            for row in rows_to_clear:
                for col in range(3, max_col + 1):
                    sheet.cell(row=row, column=col).value = None

        for i in range(80):
            # Shift the content from lines 18-40 up to lines 16-38
            for row in range(16+i*36, 39+i*36):  # Target row range
                for col in range(3, max_col + 1):  # Columns from C onwards
                    # Copy value from two rows down to the current row
                    sheet.cell(row=row, column=col).value = sheet.cell(row=row + 2, column=col).value

        # 3. Add zeros in the two empty rows (now the last two rows of each section)
        for i in range(80):  # Repeat for consistency with the above operations
            zero_start = 39 + i * 36
            zero_end = 40 + i * 36
            if zero_start > max_row: break  # Stop if beyond the last row
            for row in range(zero_start, min(zero_end + 1, max_row + 1)):
                for col in range(3, max_col + 1):
                    sheet.cell(row=row, column=col).value = 0



# Save the workbook after making the modifications
new_workbook_path = 'FTT-P-24x71_2022_S0_updated.xlsx'
workbook.save(new_workbook_path)
new_workbook_path
